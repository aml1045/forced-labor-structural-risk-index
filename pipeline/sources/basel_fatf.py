"""Basel AML Index Expert Edition -- Monetization financial-integrity signals.

Data map for the Monetization domain
`monetization/domain-a-transnational-concealment` (Product-2 lens).

Per design decision the Basel *composite* columns are DROPPED as
circular (the AML/CFT-framework domain embeds the US State Dept
Trafficking-in-Persons Report -> re-imports the trafficking outcome). This
connector instead extracts the clean SUB-COMPONENTS the scoring rules name as
the distinct Monetization financial-integrity entries (see docs/scoring-rules.md):

  basel_fatf_me_effectiveness  (PRIMARY)  -- FATF Mutual Evaluation Reports
        sub-component (0.35 of the AML/CFT-framework domain). FATF-based
        effectiveness measure; the cleanest non-circular financial-integrity
        base. Maps Domain-A defeater Z1 (effective AML enforcement) -- read in
        the RISK direction (higher = weaker AML effectiveness = MORE concealment
        risk).
  basel_tjn_fsi  (CANDIDATE second)       -- Tax Justice Network Financial
        Secrecy Index sub-component (0.15). The independent secrecy base the
        locked rule names; maps Domain-A signal s2.1 (secrecy-jurisdiction
        exposure). Higher = more secrecy = more risk.
  basel_fatf_listing_flag  (OPTIONAL binary) -- FATF grey-list (increased
        monitoring) OR black-list (call for action) standing, 0/1. A separate
        international-financial-integrity-standing signal. Higher = listed = more
        risk.

SCALE / ANCHOR (scoring rule 1 -- ABSOLUTE-anchored, not relative min-max):
  The Basel Expert Edition publishes every sub-component already rescaled to the
  Basel risk scale [0, 10] (0 = low risk, 10 = high risk). That bounded scale is
  a fixed, standards-anchored interval -- its endpoints are the absolute anchors.
  We map [0, 10] -> [0, 1] with direction=high_risk (no inversion: 10 already
  means most risk). The listing flag is anchored on its natural [0, 1] binary.

DE-CORRELATION (scoring rule 4 / rule 8 -- pending decision):
  FATF-ME and FSI are kept as DISTINCT rows and FLAGGED for the data-stage
  correlation/collinearity screen vs the general-governance backbone
  (wb_wgi_rule_of_law / v2x_rule) and vs each other. The surviving component
  count is flagged for review; this connector does NOT merge or drop.

CIRCULARITY (Domain-A circularity flags; see docs/METHODS.md):
  FSI is the index the deep-dive flagged as embedded; the Basel COMPOSITE that
  fuses secrecy+AML+TIP is excluded. These disaggregated sub-components are the
  non-circular slice the survival condition requires -- still surfaced as
  lens-only (Product-2), never promoted to the R x E composite.

LICENSE: Basel AML Index -- open with citation (Basel Institute on Governance).
  TJN FSI sub-component travels inside the Basel workbook; its standalone
  re-publication license is UNCONFIRMED -> LICENSE-PENDING flag surfaced.

Reuses iso_utils (normalize_to_iso3 handles the workbook's ISO2 codes),
standardize (AnchorSpec, anchor_scale), register (upsert_rows) -- never
re-derives them.

Run:  python -m pipeline.sources.basel_fatf
"""

from pathlib import Path
import csv

from pipeline import iso_utils, register
from pipeline.standardize import AnchorSpec, anchor_scale

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

_REPO_ROOT = Path(__file__).resolve().parent.parent.parent
_PROJECT_ROOT = _REPO_ROOT  # raw inputs resolve repo-relative under data/raw/
OUT_PATH = _REPO_ROOT / "data" / "processed" / "basel_fatf.csv"
FRAGMENT_PATH = _REPO_ROOT / "config" / "data_register.d" / "basel_fatf.csv"

# Workbook on disk (direct upload, reused as-is; under data/raw/).
WORKBOOK = (
    _PROJECT_ROOT
    / "data" / "raw"
    / "basel-aml-index-expertedition_2026-03-31.xlsx"
)
SHEET = "Expert Edition"

# Sheet layout (1-indexed columns; data starts row 5):
#   row 1 = header, row 2 = vintage date, row 3 = weight, row 4 = sub-header,
#   row 5.. = country rows.
DATA_START_ROW = 5
COL_ISO2 = 2                  # "ISO Code" (alpha-2)
COL_FATF_ME = 9              # FATF Mutual Evaluation Reports (0-10 risk)
COL_FSI = 20                # Tax Justice Network: Financial Secrecy Index (0-10 risk)
COL_GREY = 38              # FATF increased monitoring (grey list)  -> "yes"/None
COL_BLACK = 39           # FATF call for action (black list)        -> "yes"/None

BASEL_SOURCE = (
    "Basel AML Index Expert Edition 2026 (Basel Institute on Governance), "
    "FATF Mutual Evaluation sub-component"
)
FSI_SOURCE = (
    "Tax Justice Network Financial Secrecy Index 2025 "
    "(via Basel AML Index Expert Edition 2026)"
)
BASEL_LICENSE = "Open with citation (Basel Institute on Governance)"
FSI_LICENSE = (
    "Basel workbook open with citation; standalone TJN FSI re-publication "
    "license UNCONFIRMED -- RE-PUBLICATION-UNCONFIRMED"
)

# Correlation hand-off: every FI signal carries this so the correlation screen
# cannot be skipped (scoring rule 4 / rule 8).
CORR_FLAG = (
    "CORRELATION-PENDING (scoring rule 4 / rule 8): keep-distinct-vs-merge for "
    "the Monetization financial-integrity component count is the data-stage "
    "correlation/collinearity screen vs the governance backbone "
    "(wb_wgi_rule_of_law / v2x_rule) and FATF-ME vs FSI -- pending decision, "
    "flagged for review (placeholder ~r<0.6-0.7 keep / ~r>=0.85 merge)"
)
PRODUCT2_FLAG = (
    "PRODUCT-2 LENS ONLY (scoring rule 7; see docs/METHODS.md): Monetization "
    "Domain A feeds the intervention map, NOT the R x E composite"
)
CIRCULAR_FLAG = (
    "NON-CIRCULAR SLICE (circularity flags; docs/METHODS.md): disaggregated "
    "Basel sub-component, NOT the Basel composite (which embeds the TIP Report); "
    "lens-only -- do not promote to Product-1 without the data-stage "
    "demonstration"
)


def _load_workbook_rows():
    if not _HAS_OPENPYXL:
        raise RuntimeError(
            "openpyxl is required to read the Basel Expert Edition workbook "
            "(pip install openpyxl)."
        )
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Basel workbook not found: {WORKBOOK}")
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise KeyError(f"Sheet {SHEET!r} not in workbook; have {wb.sheetnames}")
    ws = wb[SHEET]
    rows = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        iso2 = row[COL_ISO2 - 1]
        country = row[0]
        if iso2 is None and country is None:
            continue
        rows.append(row)
    wb.close()
    return rows


def _extract(rows):
    """Return {iso3: value} dicts for FATF-ME, FSI, and the listing flag.

    Non-numeric / blank score cells stay MISSING (never -> 0). The listing flag
    is genuinely binary: "yes" -> 1.0, otherwise (no listing) -> 0.0 for any
    country present in the workbook; absent-from-workbook stays missing.
    """
    fatf, fsi, listing = {}, {}, {}
    unmatched = []
    for row in rows:
        iso3 = iso_utils.normalize_to_iso3(row[COL_ISO2 - 1])
        if not iso3:
            unmatched.append((row[0], row[COL_ISO2 - 1]))
            continue

        v_fatf = row[COL_FATF_ME - 1]
        if isinstance(v_fatf, (int, float)):
            fatf[iso3] = float(v_fatf)

        v_fsi = row[COL_FSI - 1]
        if isinstance(v_fsi, (int, float)):
            fsi[iso3] = float(v_fsi)

        # listing flag: present-in-workbook means we know its listing status,
        # so 0 here is a TRUE 0 (not listed), not a missing value.
        grey = str(row[COL_GREY - 1]).strip().lower() == "yes"
        black = str(row[COL_BLACK - 1]).strip().lower() == "yes"
        listing[iso3] = 1.0 if (grey or black) else 0.0

    return fatf, fsi, listing, unmatched


def run():
    countries = iso_utils.load_sample()
    rows = _load_workbook_rows()
    fatf, fsi, listing, unmatched = _extract(rows)

    columns = []        # [(output_col, ScaleResult)]
    register_rows = []

    # --- PRIMARY: FATF Mutual Evaluation effectiveness ------------------------
    spec_fatf = AnchorSpec(
        indicator="basel_fatf_me_effectiveness",
        floor=0.0, ceiling=10.0, direction="high_risk",
        unit="Basel risk scale 0-10 (FATF Mutual Evaluation effectiveness; "
             "higher = weaker AML effectiveness = more risk)",
        anchor_source=(
            "Basel AML Index Expert Edition publishes the FATF Mutual "
            "Evaluation Reports sub-component on the bounded Basel risk scale "
            "[0,10] (0 = low risk / effective AML, 10 = high risk / ineffective); "
            "the scale endpoints are the absolute anchors. Domain-A defeater Z1 "
            "read in the risk direction."
        ),
    )
    res_fatf = anchor_scale(fatf, spec_fatf, sample=countries)
    columns.append(("basel_fatf_me_effectiveness", res_fatf))
    row = res_fatf.register_row(
        source=BASEL_SOURCE,
        series_id="basel-aml-index-expertedition_2026-03-31.xlsx:FATF Mutual Evaluation Reports (col 9)",
        license=BASEL_LICENSE,
        extra_flags=[PRODUCT2_FLAG, CIRCULAR_FLAG, CORR_FLAG],
    )
    row["year_min"], row["year_max"] = 2026, 2026
    register_rows.append(row)

    # --- CANDIDATE second: TJN Financial Secrecy Index ------------------------
    spec_fsi = AnchorSpec(
        indicator="basel_tjn_fsi",
        floor=0.0, ceiling=10.0, direction="high_risk",
        unit="Basel risk scale 0-10 (TJN Financial Secrecy Index; "
             "higher = more secrecy-jurisdiction exposure = more risk)",
        anchor_source=(
            "Basel AML Index Expert Edition publishes the Tax Justice Network "
            "Financial Secrecy Index sub-component on the bounded Basel risk "
            "scale [0,10] (0 = low secrecy risk, 10 = high); endpoints are the "
            "absolute anchors. Maps Domain-A signal s2.1 (secrecy-jurisdiction "
            "exposure)."
        ),
    )
    res_fsi = anchor_scale(fsi, spec_fsi, sample=countries)
    columns.append(("basel_tjn_fsi", res_fsi))
    row = res_fsi.register_row(
        source=FSI_SOURCE,
        series_id="basel-aml-index-expertedition_2026-03-31.xlsx:Tax Justice Network Financial Secrecy Index (col 20)",
        license=FSI_LICENSE,
        extra_flags=[
            PRODUCT2_FLAG, CIRCULAR_FLAG, CORR_FLAG,
            "LICENSE-PENDING: standalone TJN FSI re-publication rights "
            "UNCONFIRMED -- confirm before public release (NOT blocking now)",
            "CANDIDATE SECOND SIGNAL: kept only if it passes the data-stage "
            "de-correlation check vs FATF-ME + the governance backbone",
        ],
    )
    row["year_min"], row["year_max"] = 2025, 2025
    register_rows.append(row)

    # --- OPTIONAL binary: FATF grey/black-list standing -----------------------
    spec_flag = AnchorSpec(
        indicator="basel_fatf_listing_flag",
        floor=0.0, ceiling=1.0, direction="high_risk",
        unit="binary 0/1 (FATF grey- or black-list standing)",
        anchor_source=(
            "FATF increased-monitoring (grey) or call-for-action (black) listing "
            "from the Basel Expert Edition sanctions block; natural binary anchor "
            "[0,1] (0 = not listed, 1 = listed). 0 is a TRUE 0 for countries in "
            "the workbook, not missing."
        ),
        # binary listing flag: only ~25 countries are listed; coverage is the
        # share of the sample WITH a known status (present in the workbook), not
        # the share listed. Keep the standard floor; this clears it easily.
    )
    res_flag = anchor_scale(listing, spec_flag, sample=countries)
    columns.append(("basel_fatf_listing_flag", res_flag))
    row = res_flag.register_row(
        source=BASEL_SOURCE,
        series_id="basel-aml-index-expertedition_2026-03-31.xlsx:FATF grey list (col 38) | FATF black list (col 39)",
        license=BASEL_LICENSE,
        extra_flags=[
            PRODUCT2_FLAG,
            "OPTIONAL listing signal (separate from the continuous FATF-ME "
            "score); screen as its own indicator in the data-stage correlation check",
        ],
    )
    row["year_min"], row["year_max"] = 2026, 2026
    register_rows.append(row)

    # --- write data/processed/basel_fatf.csv ----------------------------------
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    header = ["iso3"] + [c for c, _ in columns]
    with open(OUT_PATH, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        for iso3 in countries:
            out_row = [iso3]
            for _, result in columns:
                v = result.get(iso3)
                out_row.append("" if v is None else round(v, 4))
            w.writerow(out_row)

    # --- write per-source register fragment (NOT the master register) ---------
    register.upsert_rows(register_rows, path=str(FRAGMENT_PATH))

    for output_col, result in columns:
        print(
            f"[basel_fatf] {output_col}: dir={result.meta['direction']} "
            f"anchor={result.meta['anchor']} "
            f"coverage={result.meta['coverage_pct']:.1f}% "
            f"({result.meta['n_present']}/{result.meta['n_total']}) "
            f"below_floor={result.meta['below_floor']}"
        )
    if unmatched:
        print(f"[basel_fatf] {len(unmatched)} workbook rows unmatched to ISO3 "
              f"(dropped): {[u[1] for u in unmatched]}")
    print(f"[basel_fatf] wrote {OUT_PATH}")
    print(f"[basel_fatf] wrote fragment {FRAGMENT_PATH} "
          f"({len(register_rows)} indicator rows)")
    return columns, register_rows


if __name__ == "__main__":
    run()
